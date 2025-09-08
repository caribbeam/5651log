"""
Gelişmiş Raporlama Servisleri
5651 Loglama için kapsamlı rapor oluşturma servisleri
"""

import io
import json
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives
from django.conf import settings

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF export
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Chart generation
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    import seaborn as sns
    CHART_AVAILABLE = True
except ImportError:
    CHART_AVAILABLE = False

from .models import ReportTemplate, GeneratedReport, ReportData
from log_kayit.models import LogKayit, Company
from syslog_server.models import SyslogMessage, SyslogServer
from mirror_port.models import MirrorTraffic
from timestamp_signing.models import TimestampSignature


class ReportGenerator:
    """Rapor oluşturucu ana sınıfı"""
    
    def __init__(self, template, period_start, period_end, company):
        self.template = template
        self.period_start = period_start
        self.period_end = period_end
        self.company = company
        self.report = None
        
    def generate_report(self, format_type='pdf'):
        """Rapor oluştur"""
        # Rapor kaydı oluştur
        self.report = GeneratedReport.objects.create(
            template=self.template,
            title=f"{self.template.name} - {self.period_start.date()} - {self.period_end.date()}",
            description=f"{self.template.get_template_type_display()} raporu",
            report_format=format_type,
            period_start=self.period_start,
            period_end=self.period_end,
            status='generating',
            started_at=timezone.now()
        )
        
        try:
            # Veri toplama
            data = self._collect_data()
            
            # Rapor oluşturma
            if format_type == 'pdf':
                file_path = self._generate_pdf(data)
            elif format_type == 'excel':
                file_path = self._generate_excel(data)
            elif format_type == 'csv':
                file_path = self._generate_csv(data)
            elif format_type == 'json':
                file_path = self._generate_json(data)
            else:
                raise ValueError(f"Desteklenmeyen format: {format_type}")
            
            # Rapor güncelleme
            self.report.status = 'completed'
            self.report.completed_at = timezone.now()
            self.report.file_path = file_path
            self.report.total_records = data.get('total_records', 0)
            self.report.suspicious_records = data.get('suspicious_records', 0)
            self.report.save()
            
            return self.report
            
        except Exception as e:
            self.report.status = 'failed'
            self.report.error_message = str(e)
            self.report.completed_at = timezone.now()
            self.report.save()
            raise e
    
    def _collect_data(self):
        """Veri toplama"""
        data = {
            'company': self.company,
            'period_start': self.period_start,
            'period_end': self.period_end,
            'template': self.template,
        }
        
        # Kullanıcı logları
        if self.template.include_user_logs:
            user_logs = self._get_user_logs()
            data['user_logs'] = user_logs
            data['total_records'] = user_logs.count()
            data['suspicious_records'] = user_logs.filter(is_suspicious=True).count()
        
        # Syslog verileri
        if self.template.include_syslog_data:
            data['syslog_data'] = self._get_syslog_data()
        
        # Mirror trafik verileri
        if self.template.include_mirror_traffic:
            data['mirror_traffic'] = self._get_mirror_traffic()
        
        # Zaman damgası verileri
        if self.template.include_timestamp_data:
            data['timestamp_data'] = self._get_timestamp_data()
        
        # Güvenlik uyarıları
        if self.template.include_security_alerts:
            data['security_alerts'] = self._get_security_alerts()
        
        # İstatistikler
        data['statistics'] = self._calculate_statistics(data)
        
        # Grafikler
        if self.template.include_charts:
            data['charts'] = self._generate_charts(data)
        
        return data
    
    def _get_user_logs(self):
        """Kullanıcı loglarını al"""
        queryset = LogKayit.objects.filter(
            company=self.company,
            giris_zamani__gte=self.period_start,
            giris_zamani__lte=self.period_end
        )
        
        if self.template.include_suspicious_only:
            queryset = queryset.filter(is_suspicious=True)
        
        return queryset.order_by('-giris_zamani')
    
    def _get_syslog_data(self):
        """Syslog verilerini al"""
        return SyslogMessage.objects.filter(
            company=self.company,
            timestamp__gte=self.period_start,
            timestamp__lte=self.period_end
        ).order_by('-timestamp')
    
    def _get_mirror_traffic(self):
        """Mirror trafik verilerini al"""
        return MirrorTraffic.objects.filter(
            company=self.company,
            timestamp__gte=self.period_start,
            timestamp__lte=self.period_end
        ).order_by('-timestamp')
    
    def _get_timestamp_data(self):
        """Zaman damgası verilerini al"""
        return TimestampSignature.objects.filter(
            company=self.company,
            created_at__gte=self.period_start,
            created_at__lte=self.period_end
        ).order_by('-created_at')
    
    def _get_security_alerts(self):
        """Güvenlik uyarılarını al"""
        # Bu kısım security_alerts modülüne bağlı olacak
        return []
    
    def _calculate_statistics(self, data):
        """İstatistikleri hesapla"""
        stats = {}
        
        if 'user_logs' in data:
            user_logs = data['user_logs']
            stats['user_logs'] = {
                'total': user_logs.count(),
                'suspicious': user_logs.filter(is_suspicious=True).count(),
                'daily_average': user_logs.count() / max(1, (self.period_end - self.period_start).days),
                'top_users': user_logs.values('ad_soyad').annotate(
                    count=Count('id')
                ).order_by('-count')[:10]
            }
        
        if 'syslog_data' in data:
            syslog_data = data['syslog_data']
            stats['syslog'] = {
                'total': syslog_data.count(),
                'by_facility': syslog_data.values('facility').annotate(
                    count=Count('id')
                ).order_by('-count'),
                'by_priority': syslog_data.values('priority').annotate(
                    count=Count('id')
                ).order_by('-count')
            }
        
        return stats
    
    def _generate_charts(self, data):
        """Grafikler oluştur"""
        charts = {}
        
        if not CHART_AVAILABLE:
            return charts
        
        # Günlük giriş grafiği
        if 'user_logs' in data:
            charts['daily_logins'] = self._create_daily_logins_chart(data['user_logs'])
        
        # Syslog priority dağılımı
        if 'syslog_data' in data:
            charts['syslog_priority'] = self._create_syslog_priority_chart(data['syslog_data'])
        
        return charts
    
    def _create_daily_logins_chart(self, user_logs):
        """Günlük giriş grafiği oluştur"""
        if not CHART_AVAILABLE:
            return None
        
        # Günlük verileri hazırla
        daily_data = {}
        current_date = self.period_start.date()
        end_date = self.period_end.date()
        
        while current_date <= end_date:
            daily_data[current_date] = user_logs.filter(
                giris_zamani__date=current_date
            ).count()
            current_date += timedelta(days=1)
        
        # Grafik oluştur
        fig, ax = plt.subplots(figsize=(12, 6))
        dates = list(daily_data.keys())
        counts = list(daily_data.values())
        
        ax.plot(dates, counts, marker='o', linewidth=2, markersize=6)
        ax.set_title('Günlük Kullanıcı Girişleri', fontsize=14, fontweight='bold')
        ax.set_xlabel('Tarih', fontsize=12)
        ax.set_ylabel('Giriş Sayısı', fontsize=12)
        ax.grid(True, alpha=0.3)
        
        # Tarih formatı
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))
        ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
        plt.xticks(rotation=45)
        
        plt.tight_layout()
        
        # Base64'e çevir
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        chart_data = buffer.getvalue()
        plt.close()
        
        return chart_data
    
    def _create_syslog_priority_chart(self, syslog_data):
        """Syslog priority dağılım grafiği oluştur"""
        if not CHART_AVAILABLE:
            return None
        
        # Priority verilerini hazırla
        priority_data = syslog_data.values('priority').annotate(
            count=Count('id')
        ).order_by('priority')
        
        priorities = []
        counts = []
        for item in priority_data:
            priorities.append(f"Priority {item['priority']}")
            counts.append(item['count'])
        
        # Grafik oluştur
        fig, ax = plt.subplots(figsize=(10, 6))
        colors = ['#ff4444', '#ff8800', '#ffaa00', '#ffdd00', '#88ff88', '#44ff44', '#00aaff', '#0088ff']
        
        ax.pie(counts, labels=priorities, autopct='%1.1f%%', colors=colors[:len(priorities)])
        ax.set_title('Syslog Priority Dağılımı', fontsize=14, fontweight='bold')
        
        plt.tight_layout()
        
        # Base64'e çevir
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        chart_data = buffer.getvalue()
        plt.close()
        
        return chart_data
    
    def _generate_pdf(self, data):
        """PDF raporu oluştur"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab kütüphanesi yüklü değil")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Stil tanımları
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Başlık
        story.append(Paragraph(f"{self.template.name}", title_style))
        story.append(Spacer(1, 12))
        
        # Rapor bilgileri
        info_data = [
            ['Şirket:', self.company.name],
            ['Dönem:', f"{self.period_start.date()} - {self.period_end.date()}"],
            ['Oluşturulma:', timezone.now().strftime('%d.%m.%Y %H:%M')],
            ['Tip:', self.template.get_template_type_display()],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # İstatistikler
        if 'statistics' in data and self.template.include_statistics:
            story.append(Paragraph("İstatistikler", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            stats = data['statistics']
            if 'user_logs' in stats:
                user_stats = stats['user_logs']
                stats_data = [
                    ['Toplam Giriş:', str(user_stats['total'])],
                    ['Şüpheli Giriş:', str(user_stats['suspicious'])],
                    ['Günlük Ortalama:', f"{user_stats['daily_average']:.1f}"],
                ]
                
                stats_table = Table(stats_data, colWidths=[2*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(stats_table)
                story.append(Spacer(1, 20))
        
        # Detaylar
        if 'user_logs' in data and self.template.include_details:
            story.append(Paragraph("Kullanıcı Giriş Detayları", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Tablo başlıkları
            table_data = [['TC No', 'Ad Soyad', 'IP Adresi', 'Giriş Zamanı', 'Şüpheli']]
            
            # İlk 50 kayıt
            for log in data['user_logs'][:50]:
                table_data.append([
                    log.tc_no or '-',
                    log.ad_soyad or '-',
                    log.ip_adresi or '-',
                    log.giris_zamani.strftime('%d.%m.%Y %H:%M'),
                    'Evet' if log.is_suspicious else 'Hayır'
                ])
            
            if len(data['user_logs']) > 50:
                table_data.append(['...', f'Toplam {len(data["user_logs"])} kayıt', '', '', ''])
            
            detail_table = Table(table_data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1.2*inch, 0.8*inch])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(detail_table)
        
        # PDF oluştur
        doc.build(story)
        buffer.seek(0)
        
        # Dosyayı kaydet
        filename = f"report_{self.report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = f"{settings.MEDIA_ROOT}/reports/{filename}"
        
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        
        return file_path
    
    def _generate_excel(self, data):
        """Excel raporu oluştur"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl kütüphanesi yüklü değil")
        
        wb = openpyxl.Workbook()
        
        # Ana sayfa
        ws = wb.active
        ws.title = "Rapor Özeti"
        
        # Başlık
        ws['A1'] = self.template.name
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Rapor bilgileri
        ws['A3'] = 'Şirket:'
        ws['B3'] = self.company.name
        ws['A4'] = 'Dönem:'
        ws['B4'] = f"{self.period_start.date()} - {self.period_end.date()}"
        ws['A5'] = 'Oluşturulma:'
        ws['B5'] = timezone.now().strftime('%d.%m.%Y %H:%M')
        
        # İstatistikler
        if 'statistics' in data and self.template.include_statistics:
            row = 7
            ws[f'A{row}'] = 'İstatistikler'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            stats = data['statistics']
            if 'user_logs' in stats:
                user_stats = stats['user_logs']
                ws[f'A{row}'] = 'Toplam Giriş:'
                ws[f'B{row}'] = user_stats['total']
                row += 1
                ws[f'A{row}'] = 'Şüpheli Giriş:'
                ws[f'B{row}'] = user_stats['suspicious']
                row += 1
                ws[f'A{row}'] = 'Günlük Ortalama:'
                ws[f'B{row}'] = round(user_stats['daily_average'], 1)
                row += 2
        
        # Kullanıcı logları sayfası
        if 'user_logs' in data and self.template.include_details:
            ws_logs = wb.create_sheet("Kullanıcı Logları")
            
            # Başlıklar
            headers = ['TC No', 'Ad Soyad', 'IP Adresi', 'MAC Adresi', 'Giriş Zamanı', 'Şüpheli']
            for col, header in enumerate(headers, 1):
                cell = ws_logs.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Veriler
            for row, log in enumerate(data['user_logs'], 2):
                ws_logs.cell(row=row, column=1, value=log.tc_no or '-')
                ws_logs.cell(row=row, column=2, value=log.ad_soyad or '-')
                ws_logs.cell(row=row, column=3, value=log.ip_adresi or '-')
                ws_logs.cell(row=row, column=4, value=log.mac_adresi or '-')
                ws_logs.cell(row=row, column=5, value=log.giris_zamani.strftime('%d.%m.%Y %H:%M'))
                ws_logs.cell(row=row, column=6, value='Evet' if log.is_suspicious else 'Hayır')
            
            # Sütun genişlikleri
            for col in range(1, len(headers) + 1):
                ws_logs.column_dimensions[get_column_letter(col)].width = 15
        
        # Dosyayı kaydet
        filename = f"report_{self.report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = f"{settings.MEDIA_ROOT}/reports/{filename}"
        
        wb.save(file_path)
        return file_path
    
    def _generate_csv(self, data):
        """CSV raporu oluştur"""
        import csv
        
        filename = f"report_{self.report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = f"{settings.MEDIA_ROOT}/reports/{filename}"
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Başlık
            writer.writerow([self.template.name])
            writer.writerow([f"Şirket: {self.company.name}"])
            writer.writerow([f"Dönem: {self.period_start.date()} - {self.period_end.date()}"])
            writer.writerow([f"Oluşturulma: {timezone.now().strftime('%d.%m.%Y %H:%M')}"])
            writer.writerow([])
            
            # Kullanıcı logları
            if 'user_logs' in data and self.template.include_details:
                writer.writerow(['Kullanıcı Logları'])
                writer.writerow(['TC No', 'Ad Soyad', 'IP Adresi', 'MAC Adresi', 'Giriş Zamanı', 'Şüpheli'])
                
                for log in data['user_logs']:
                    writer.writerow([
                        log.tc_no or '-',
                        log.ad_soyad or '-',
                        log.ip_adresi or '-',
                        log.mac_adresi or '-',
                        log.giris_zamani.strftime('%d.%m.%Y %H:%M'),
                        'Evet' if log.is_suspicious else 'Hayır'
                    ])
        
        return file_path
    
    def _generate_json(self, data):
        """JSON raporu oluştur"""
        filename = f"report_{self.report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path = f"{settings.MEDIA_ROOT}/reports/{filename}"
        
        # Veriyi JSON'a uygun hale getir
        json_data = {
            'template': {
                'name': self.template.name,
                'type': self.template.get_template_type_display(),
            },
            'company': {
                'name': self.company.name,
                'slug': self.company.slug,
            },
            'period': {
                'start': self.period_start.isoformat(),
                'end': self.period_end.isoformat(),
            },
            'generated_at': timezone.now().isoformat(),
            'statistics': data.get('statistics', {}),
        }
        
        # Kullanıcı logları
        if 'user_logs' in data and self.template.include_details:
            json_data['user_logs'] = []
            for log in data['user_logs']:
                json_data['user_logs'].append({
                    'tc_no': log.tc_no,
                    'ad_soyad': log.ad_soyad,
                    'ip_adresi': log.ip_adresi,
                    'mac_adresi': log.mac_adresi,
                    'giris_zamani': log.giris_zamani.isoformat(),
                    'is_suspicious': log.is_suspicious,
                })
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        return file_path


class ReportScheduler:
    """Rapor zamanlayıcı"""
    
    def __init__(self):
        pass
    
    def run_scheduled_reports(self):
        """Zamanlanmış raporları çalıştır"""
        from .models import ReportSchedule
        
        now = timezone.now()
        schedules = ReportSchedule.objects.filter(
            is_active=True,
            next_run__lte=now
        )
        
        for schedule in schedules:
            try:
                self._run_schedule(schedule)
            except Exception as e:
                # Hata logla
                print(f"Rapor zamanlaması hatası: {e}")
    
    def _run_schedule(self, schedule):
        """Tek bir zamanlamayı çalıştır"""
        template = schedule.template
        
        # Dönem hesapla
        period_start, period_end = self._calculate_period(schedule)
        
        # Rapor oluştur
        generator = ReportGenerator(template, period_start, period_end, template.company)
        report = generator.generate_report('pdf')
        
        # Sonraki çalışma zamanını hesapla
        schedule.last_run = now
        schedule.next_run = self._calculate_next_run(schedule)
        schedule.save()
        
        return report
    
    def _calculate_period(self, schedule):
        """Dönem hesapla"""
        now = timezone.now()
        
        if schedule.frequency == 'daily':
            period_start = now - timedelta(days=1)
            period_end = now
        elif schedule.frequency == 'weekly':
            period_start = now - timedelta(weeks=1)
            period_end = now
        elif schedule.frequency == 'monthly':
            period_start = now - timedelta(days=30)
            period_end = now
        else:
            period_start = now - timedelta(days=1)
            period_end = now
        
        return period_start, period_end
    
    def _calculate_next_run(self, schedule):
        """Sonraki çalışma zamanını hesapla"""
        now = timezone.now()
        
        if schedule.frequency == 'daily':
            next_run = now + timedelta(days=1)
        elif schedule.frequency == 'weekly':
            next_run = now + timedelta(weeks=1)
        elif schedule.frequency == 'monthly':
            next_run = now + timedelta(days=30)
        else:
            next_run = now + timedelta(days=1)
        
        # Saati ayarla
        next_run = next_run.replace(
            hour=schedule.run_time.hour,
            minute=schedule.run_time.minute,
            second=0,
            microsecond=0
        )
        
        return next_run
